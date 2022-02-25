"""Datasource builder for data from British Petroleum."""

from typing import Any, Dict, List, Optional

import openpyxl
import pandas as pd

from shifter_pandas import standardize_property
from shifter_pandas.wikidata_ import WikidataDatasource

UNITS_ENERGY = [
    "Exajoules",
    "Petajoules",
    "Exajoules (input-equivalent)",
    "Exajoules (input-equivalent)*",
    "Terawatt-hours",
]
UNITS_VOLUME = [
    "Thousand barrels of oil equivalent per day",
    "Thousand barrels daily",
    "Billion cubic feet per day",
    "Trillion cubic metres",
    "Thousand million barrels",
    "Thousand barrels daily*",
    "Billion cubic metres",
]
UNITS_MASS = [
    "Thousand tonnes1",
    "Million tonnes",
    "Million tonnes",
    "Million tonnes of carbon dioxide",
    "Thousand tonnes",
    "Thousand tonnes of Lithium content",
]
UNITS_ENERGY_CAPITA = [
    "Gigajoule per capita",
]


class BPDatasource:
    """Datasource builder for data from British Petroleum."""

    def __init__(self, file_name: str) -> None:
        """Initialize the datasource builder."""
        self.xlsx = openpyxl.load_workbook(file_name)
        self.wdds = WikidataDatasource()
        self.wdds.set_alias("World", "World", "Q16502", "World")

    def metadata(self) -> List[Dict[str, Any]]:
        """Get the metadata."""
        metadata: List[Dict[str, Any]] = []
        for type_index, type_value in enumerate(self.xlsx.sheetnames):
            if type_value.endswith(" - TWh"):
                type_value = type_value[:-6]
            if type_value.endswith(" - EJ"):
                type_value = type_value[:-5]
            if type_value.endswith(" - PJ"):
                type_value = type_value[:-5]

            if isinstance(self.xlsx.worksheets[type_index].cell(3, 2).value, int):
                # Year
                years = []
                index = 2
                while True:
                    value = self.xlsx.worksheets[type_index].cell(3, index).value
                    if self.xlsx.worksheets[type_index].cell(2, index).value is not None:
                        break
                    years.append({"label": value, "index": index})
                    index += 1

                # Country
                regions = []
                index = 5
                nb_empty_cells = 0
                while True:
                    value = self.xlsx.worksheets[type_index].cell(index, 1).value
                    if value is not None:
                        nb_empty_cells = 0
                        regions.append({"label": value, "index": index})
                    nb_empty_cells += 1
                    if nb_empty_cells > 5:
                        break
                    index += 1
                metadata.append(
                    {
                        "label": type_value,
                        "index": type_index,
                        "unit": self.xlsx.worksheets[type_index].cell(3, 1).value.strip(),
                        "years": years,
                        "regions": regions,
                    }
                )

        return metadata

    def datasource(
        self,
        types_filter: Optional[List[str]] = None,
        regions_filter: Optional[List[str]] = None,
        units_filter: Optional[List[str]] = None,
        years_filter: Optional[List[str]] = None,
        years_factor: Optional[int] = None,
        coherent_units: bool = True,
        wikidata_id: bool = False,
        wikidata_type: bool = False,
        wikidata_name: bool = False,
        wikidata_properties: Optional[List[str]] = None,
    ) -> pd.DataFrame:
        """Get the Datasource as DataFrame."""

        if wikidata_properties is None:
            wikidata_properties = []
        wikidata = wikidata_id or wikidata_name or wikidata_properties

        columns = ["Value", "Type", "Unit", "TypeUnit", "Year", "Region"]
        if wikidata:
            if wikidata_id:
                columns.append("WikidataId")
            if wikidata_name:
                columns.append("WikidataName")
            if wikidata_type:
                columns.append("WikidataType")
            for wikidata_property in wikidata_properties:
                columns.append(
                    f"Wikidata{standardize_property(self.wdds.get_property_name(wikidata_property))}"
                )
        data_frame = pd.DataFrame(columns=columns)
        for type_ in self.metadata():
            type_index = type_["index"]
            type_label = type_["label"]
            years = type_["years"]
            regions = type_["regions"]

            if types_filter is not None and type_label not in types_filter:
                continue

            for year in years:
                if years_filter is not None and year["label"] not in years_filter:
                    continue
                if years_factor is not None and year["label"] % years_factor != 0:
                    continue
                for region in regions:
                    if regions_filter is not None and region["label"] not in regions_filter:
                        continue
                    value = self.xlsx.worksheets[type_index].cell(region["index"], year["index"]).value
                    if not isinstance(value, int) and not isinstance(value, float):
                        continue

                    unit = self.xlsx.worksheets[type_index].cell(3, 1).value

                    if units_filter is not None and unit not in units_filter:
                        continue
                    if coherent_units:
                        if unit == "Terawatt-hours":
                            value = value * 3.6
                            unit = "Petajoules"
                        if unit == "Exajoules":
                            value = value * 1000
                            unit = "Petajoules"
                        if unit == "Exajoules (input-equivalent)":
                            value = value * 1000
                            unit = "Petajoules (input-equivalent)"

                    element = {
                        "Value": value,
                        "Year": year["label"],
                        "Region": region["label"],
                        "Type": type_label,
                        "Unit": unit,
                        "TypeUnit": f"{type_label} [{unit}]",
                    }
                    if wikidata:
                        region_label = region["label"]
                        if region_label.startswith("Total "):
                            region_label = region_label[6:]
                        element_id = self.wdds.get_region(region_label)
                        element["WikidataType"] = element_id['type'] if element_id else None
                        element.update(
                            self.wdds.get_item(
                                element_id["id"] if element_id else None,
                                with_name=wikidata_name,
                                with_id=wikidata_id,
                                properties=wikidata_properties,
                                prefix="Wikidata",
                            )
                        )
                    data_frame = data_frame.append(element, ignore_index=True)

        return data_frame
